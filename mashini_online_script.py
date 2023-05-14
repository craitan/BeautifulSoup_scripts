from bs4 import BeautifulSoup
import requests
import openpyxl

# def get_categories_links():
#     url = "https://www.onlinemashini.bg/"
#     response = requests.get(url)
#
#     doc = BeautifulSoup(response.content, 'html.parser')
#
#     links = doc.find_all('a', {'class': 'submenu active'})
#
#     list_of_links = []
#
#     for link in links:
#         split_link_1 = str(link).split(' ')
#         split_link_2 = str(split_link_1[3]).split('"')
#         list_of_links.append('https://www.onlinemashini.bg' + str(split_link_2[1]))
#
#     return list_of_links


def get_products_links():
    url = 'https://www.onlinemashini.bg/cat/108/gradinska-tehnika/192/drobilki-za-kloni'

    """Creating list to save the products links"""
    list_of_products_link = []

    response = requests.get(url)
    doc = BeautifulSoup(response.content, 'html.parser')

    """Looking if there are pages"""
    get_pages = doc.find('ul', {'class': 'pagination'})

    """If there are pages we want to get the last page number using slicing"""
    if get_pages != None:
        last_page_link = get_pages.find_all('a')[-1]  # get the last <a> tag
        last_page_href = last_page_link['href']  # extract the "href" attribute
        last_page_number = int(last_page_href.split('/')[-1])  # get the last segment of the URL

        """We create a loop with range 1 and the last page number"""
        for page in range(1, last_page_number + 1):

            """This peace of code is to break the loop if we want to"""
            if len(list_of_products_link) == 100:
                break

            """Creating new url with the page number"""
            paged_url = f'{url}/{page}'
            response = requests.get(paged_url).text
            doc = BeautifulSoup(response, 'html.parser')

            """Creating a loop that will create the links of the products and add them to the list_of_products_link"""
            for link in doc.find_all('div', {'class': 'col-8'}):
                split_link = ""
                if '<a class="name" href="/' in str(link):
                    split_link = str(link).split('<a class="name" href="/')[1].split('</a>')[0].split('">')[0]
                else:
                    continue

                """Creating new url for the given product and add it to the list_of_products_link"""
                product_link = f'https://www.onlinemashini.bg/{str(split_link)}'
                list_of_products_link.append(product_link)

                """This peace of code is to break the loop if we want to"""
                if len(list_of_products_link) == 100:
                    break

        """If there is only one page with products"""
    elif get_pages == None and len(doc.find_all('div', {'class': 'col-8'})) > 0:

        for link in doc.find_all('div', {'class': 'col-8'}):
            split_link = ""
            if '<a class="name" href="/' in str(link):
                split_link = str(link).split('<a class="name" href="/')[1].split('</a>')[0].split('">')[0]
            else:
                continue
            product_link = f'https://www.onlinemashini.bg/{str(split_link)}'
            list_of_products_link.append(product_link)

            """This peace of code is to break the loop if we want to"""
            if len(list_of_products_link) == 100:
                break
        """If there are no products and pages pass"""
    else:
        pass

    return list_of_products_link


def get_product_info():
    """Create a dict of products as key the product name and value is a list
    with the price, characteristics and link to the product image."""
    products = {}

    """The links of the products that is returned from the get_products_links function."""
    list_of_products_urls = get_products_links()

    """Looping through links"""
    for url in list_of_products_urls:
        response = requests.get(url)
        doc = BeautifulSoup(response.content, 'html.parser')
        name = doc.find(class_='product-name').text.strip()

        """Some products have a discount and with this logic we check that and take the 
        corresponding price to add it to the xlsx file"""
        if doc.find('div', class_='price mb-3 mt-3'):
            price = (doc.find(class_='price mb-3 mt-3').text.strip('лв.'))
        else:
            price = (doc.find(class_='price mb-2 mt-1').text.strip('лв.'))

        """In some prices there is a comma with this logic we remove it so that it does not give errors and 
        divide it by 100 so that we can then format it later and add it to the xlsx file"""
        price = price.replace(',', '')
        price = int(price) / 100

        """With this logic we take the characteristics of the product"""
        product_info = doc.find('div', {'class': 'decs details'})
        product_characteristics = product_info.get_text(strip=True, separator="\n")

        """with this logic, we take the link to the photo of the product, if the 
        product does not have a photo, we add the following text 'No image'"""
        if doc.find('img', class_='b-r-4 product-main-image'):
            product_img = doc.find('img', class_='b-r-4 product-main-image')
            img_link = product_img["src"]
        else:
            img_link = 'No image'

        """We add the products with their price formatted to the second 
        character, their characteristics and a link to the photo of the product"""
        if name not in products:
            products[name] = [f'{price:.2f}', str(product_characteristics), img_link]

    return products


def create_xlsx_file_with_data():
    '''Create an xlsx file'''
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    """Creates the columns"""
    worksheet['A1'] = 'Product Name'
    worksheet['B1'] = 'Price'
    worksheet['C1'] = 'Info'
    worksheet['D1'] = 'Product Img link'

    """Loop through the product dictionary and add their names, prices,
    features and a link to the photo and add them on a new line"""
    row = 2
    for product_name, data in get_product_info().items():
        worksheet.cell(row=row, column=1, value=product_name)
        worksheet.cell(row=row, column=2, value=data[0])
        worksheet.cell(row=row, column=3, value=data[1])
        worksheet.cell(row=row, column=4, value=data[2])

        row += 1

    """Save the file"""
    workbook.save('mashini_online_bg.xlsx')


create_xlsx_file_with_data()

